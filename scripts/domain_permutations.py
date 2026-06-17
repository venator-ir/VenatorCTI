# domain_permutations.py
import dns.resolver
import requests
import pandas as pd
import base64
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random
import idna
import os
from datetime import datetime
import json
import pathlib
from requests.exceptions import Timeout, HTTPError, RequestException
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

load_dotenv()

# ---------------- Paths ----------------
BASE_DIR = pathlib.Path(os.getcwd()).resolve()        # ./outputs when launched by Streamlit
DATA_DIR = (BASE_DIR / "../data").resolve()           # ../data

# ---------------- API keys / config ----------------
VT_API_KEY = os.getenv("VT_API_KEY")
RAPIDAPI_KEY = os.getenv("RAPIDAPI_KEY")
URLSCAN_API_KEY = os.getenv("URLSCAN_API_KEY")
client_name = os.getenv("CLIENT_NAME", "company")

VT_RATE_LIMIT = 2  # requests per minute
VT_TIMEOUT = 30    # Timeout for API requests in seconds
MAX_RETRIES = 2
RAPIDAPI_HOST = 'domainr.p.rapidapi.com'
WHOIS55_HOST = 'whois55.p.rapidapi.com'

# Only keep these TLDs from DNSTwister permutations
KEEP_TLDS = (".com", ".net", ".org")

# ---------------- Country mapping ----------------
country_mapping = {
    "DE": "Germany", "US": "United States", "CA": "Canada", "GB": "United Kingdom",
    "FR": "France", "JP": "Japan", "AU": "Australia", "CN": "China", "IN": "India",
    "IT": "Italy", "ES": "Spain", "VT": "Vietnam", "NL": "Netherlands", "BR": "Brazil",
    "RU": "Russia", "KR": "South Korea", "MX": "Mexico", "ZA": "South Africa",
    "SE": "Sweden", "CH": "Switzerland", "SG": "Singapore", "HK": "Hong Kong"
}

# ---------------- Rate limiter ----------------
def rate_limiter():
    time.sleep(60 / VT_RATE_LIMIT)

# ---------------- DNSTwister helpers ----------------
def get_domain_hex(domain):
    encoded_domain = idna.encode(domain).decode('utf-8')
    url = f"https://dnstwister.report/api/to_hex/{encoded_domain}"
    response = requests.get(url, timeout=20)
    response.raise_for_status()
    return response.json()['domain_as_hexadecimal']

def get_domain_permutations(domain_hex):
    url = f"https://dnstwister.report/api/fuzz/{domain_hex}"
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return response.json().get('fuzzy_domains', [])

def resolve_a_record(domain):
    """Return first DNS A record IP, or 'No DNS A Record'."""
    try:
        resolver = dns.resolver.Resolver()
        resolver.nameservers = ["1.1.1.1", "8.8.8.8"]
        resolver.lifetime = 5
        resolver.timeout = 3

        answers = resolver.resolve(domain, "A")

        ips = [r.to_text() for r in answers]
        return ips[0] if ips else "No DNS A Record"

    except Exception:
        return "No DNS A Record"

# ---------------- Misc helpers ----------------
def fetch_data_from_url(url):
    response = requests.get(url, timeout=20)
    response.raise_for_status()
    return response.json()

def get_country_from_ip(ip):
    try:
        url = f"https://ipinfo.io/{ip}/json"
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        country_code = response.json().get('country', 'No DNS A Record')
        return country_mapping.get(country_code, country_code)
    except Exception:
        return 'No DNS A Record'

def get_main_domain(domain):
    parts = domain.split('.')
    if len(parts) > 2:
        return '.'.join(parts[-2:])
    return domain

def get_virustotal_url_status(domain):
    if not VT_API_KEY:
        return "Not Checked"
    main_domain = get_main_domain(domain)
    for retry_count in range(MAX_RETRIES):
        try:
            url_id = base64.urlsafe_b64encode(main_domain.encode()).decode().strip("=")
            vt_url = f"https://www.virustotal.com/api/v3/urls/{url_id}"
            headers = {'x-apikey': VT_API_KEY}
            response = requests.get(vt_url, headers=headers, timeout=VT_TIMEOUT)
            rate_limiter()

            if response.status_code in (204, 429):
                time.sleep(60)
                continue

            response.raise_for_status()
            result = response.json()
            last_analysis_stats = result.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            if last_analysis_stats.get("malicious", 0) > 0:
                return "Possibly Malicious"
            return "Clean"
        except requests.exceptions.Timeout:
            return "Timeout"
        except requests.exceptions.RequestException:
            if retry_count >= MAX_RETRIES - 1:
                return "Clean"
    return "Clean"

def get_virustotal_ip_status(ip):
    if not VT_API_KEY or ip in (None, "", False, "No DNS A Record"):
        return "No DNS A Record" if ip == "No DNS A Record" else "Not Checked"
    for retry_count in range(MAX_RETRIES):
        try:
            vt_url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
            headers = {'x-apikey': VT_API_KEY}
            response = requests.get(vt_url, headers=headers, timeout=VT_TIMEOUT)
            rate_limiter()

            if response.status_code in (204, 429):
                time.sleep(60)
                continue

            response.raise_for_status()
            result = response.json()
            last_analysis_stats = result.get("data", {}).get("attributes", {}).get("last_analysis_stats", {})
            if last_analysis_stats.get("malicious", 0) > 0:
                return "Possibly Malicious"
            return "Clean"
        except requests.exceptions.Timeout:
            return "Timeout"
        except requests.exceptions.RequestException:
            if retry_count >= MAX_RETRIES - 1:
                return "Clean"
    return "Clean"

def get_domain_availability_status(status):
    if not status:
        return "Unknown status"
    s = status.lower()
    if "marketed" in s:
        return "Explicitly marketed as for sale via the aftermarket."
    if "reserved" in s:
        return "Explicitly reserved by ICANN, the registry, or another party."
    if "inactive" in s:
        return "Available for new registration."
    if "parked" in s:
        return "Active and parked, possibly available via the aftermarket."
    if "active" in s:
        return "Registered, but possibly available via the aftermarket. - check screenshot if domain is parked, active, or offline."

    descriptions = {
        "active": "Registered, but possibly available via the aftermarket. - check screenshot if domain is parked, active, or offline.",
        "undelegated": "The domain is not present in DNS.",
        "available": "The domain is For Sale.",
        "unavailable": "The domain is not available for registration.",
        "unknown": "Unknown status.",
        "deleted": "The domain has been deleted and is not currently active.",
        "on hold": "The domain registration is on hold, usually due to administrative reasons.",
        "premium": "Premium domain name for sale by the registry.",
        "marketed priced active": "The domain is For Sale- An aftermarket domain with an explicit price.",
        "active parked": "Registered and parked, but possibly available via the aftermarket.",
        "undelegated inactive": "The domain is not present in DNS."
    }
    return descriptions.get(status, "Unknown status")

def check_domain_availability_status(domain):
    if not RAPIDAPI_KEY:
        return "Unknown status"
    main_domain = '.'.join(domain.split('.')[-2:])
    url = f"https://domainr.p.rapidapi.com/v2/status?domain={main_domain}"
    headers = {
        'x-rapidapi-host': RAPIDAPI_HOST,
        'x-rapidapi-key': RAPIDAPI_KEY
    }
    response = requests.get(url, headers=headers, timeout=20)
    response.raise_for_status()
    status = response.json()['status'][0]['status']
    return get_domain_availability_status(status)

def get_screenshot_url(domain):
    if not URLSCAN_API_KEY:
        return None
    headers = {'API-Key': URLSCAN_API_KEY, 'Content-Type': 'application/json'}
    data = {"url": f"http://{domain}", "visibility": "public"}
    response = requests.post('https://urlscan.io/api/v1/scan/', headers=headers, data=json.dumps(data), timeout=20)
    response_data = response.json()
    uuid = response_data.get("uuid")
    if uuid:
        return f"https://urlscan.io/screenshots/{uuid}.png"
    return None

def get_last_dns_change(domain):
    if not RAPIDAPI_KEY:
        return 'N/A'
    url = f"https://{WHOIS55_HOST}/api/v1/whois?domain={domain}"
    headers = {
        'x-rapidapi-key': RAPIDAPI_KEY,
        'x-rapidapi-host': WHOIS55_HOST
    }
    retries = 3
    timeout = 10  # seconds

    for attempt in range(retries):
        try:
            response = requests.get(url, headers=headers, timeout=timeout)
            response.raise_for_status()
            if response.status_code == 200:
                whois_data = response.json()
                updated_date_str = whois_data.get('parsed', {}).get('Updated Date', 'N/A')
                if updated_date_str != 'N/A':
                    updated_date = datetime.strptime(updated_date_str, '%Y-%m-%dT%H:%M:%SZ')
                    return updated_date.strftime('%Y-%m-%d')
                return 'N/A'
        except Timeout:
            print(f"Timeout occurred for {domain}. Retrying...")
        except HTTPError as http_err:
            print(f"HTTP error occurred for {domain}: {http_err}")
            break
        except RequestException as req_err:
            print(f"Request exception occurred for {domain}: {req_err}")
            break
        except Exception as e:
            print(f"Unexpected error occurred for {domain}: {e}")
            break
        time.sleep(2)  # Wait before retrying

    return 'N/A'

# ---------------- Core processing ----------------
def process_domain_permutation(perm, base_subdomain, original_domain, original_tld):
    """
    Returns a dict row for a single permutation OR None if filtered out / error.
    """
    try:
        dom = (perm.get('domain') or '').strip().lower()
        if not dom:
            return None

        # ✅ Only keep .com/.net/.org permutations, and not the original base domain
        if not dom.endswith(KEEP_TLDS):
            return None
        if dom == original_domain.lower():
            return None

        # (rest of your enrichment stays the same)
        mx_data = fetch_data_from_url(perm['has_mx_url'])
        if not mx_data.get('mx', False):
            return None

        ip_value = resolve_a_record(dom)
        parked_data = fetch_data_from_url(perm['parked_score_url'])
        country = get_country_from_ip(ip_value) if ip_value != 'No DNS A Record' else 'No DNS A Record'
        domain_status = get_virustotal_url_status(dom)
        ip_status = get_virustotal_ip_status(ip_value) if ip_value != 'No DNS A Record' else 'No DNS A Record'
        availability_status = check_domain_availability_status(dom) if ip_value != 'No DNS A Record' else 'The domain is registered, but the website is Offline.'
        screenshot_url = get_screenshot_url(dom) if ip_value != 'No DNS A Record' else 'No DNS A Record'

        notes = ""
        if domain_status == "Possibly Malicious" and availability_status.startswith("The website is active"):
            notes = "Current Malicious Domain Behavior"
        elif domain_status == "Clean" and ip_status == "Possibly Malicious" and availability_status.startswith("The website is active"):
            notes = "Current Malicious IP Behavior"
        elif domain_status == "Possibly Malicious" and availability_status in [
            "The domain is For Sale.", "The domain is Parked.", "The domain is registered, but the website is Offline."
        ]:
            notes = "Past Malicious Domain Behavior"
        elif domain_status == "Clean" and ip_status == "Possibly Malicious" and availability_status in [
            "The domain is For Sale.", "The domain is Parked.", "The domain is registered, but the website is Offline."
        ]:
            notes = "Past Malicious IP Behavior"
        else:
            notes = "No Malicious Behavior"

        return {
            'Domain permutation': dom,
            'IP': ip_value,
            'Country': country,
            'Redirects To': parked_data.get('redirects_to', 'N/A'),
            'Domain Reputation Status': domain_status,
            'IP Reputation Status': ip_status,
            'Domain Availability Status': availability_status,
            'Website Screenshot': screenshot_url,
            'notes': notes,
            'Last DNS Change': get_last_dns_change(dom),
            'Original Domain': base_subdomain,  # used for summary/coloring
            'New/Changed': 'No'  # will be updated later
        }
    except Exception as e:
        print(f"Error processing permutation {perm.get('domain')}: {e}")
    return None

def process_subdomain(subdomain, tlds, original_tld):
    all_data = []
    permutation_count = 0

    # ✅ Deduplicate tlds + [original_tld] so we don't call DNSTwister twice for same TLD
    unique_tlds = []
    for t in tlds + [original_tld]:
        if t not in unique_tlds:
            unique_tlds.append(t)

    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = []

        for tld in unique_tlds:
            domain = f"{subdomain}.{tld}"
            try:
                domain_hex = get_domain_hex(domain)
                if not domain_hex:
                    continue
                permutations = get_domain_permutations(domain_hex)

                # ✅ Pre-filter permutations to just .com/.net/.org and not the original domain being queried
                filtered_perms = []
                for perm in permutations:
                    d = (perm.get('domain') or '').strip().lower()
                    if not d:
                        continue
                    if not d.endswith(KEEP_TLDS):
                        continue
                    if d == domain.lower():  # avoid processing the same base domain
                        continue
                    filtered_perms.append(perm)

                permutation_count += len(filtered_perms)

                for perm in filtered_perms:
                    futures.append(executor.submit(
                        process_domain_permutation,
                        perm,
                        f"{subdomain}.{original_tld}",  # base subdomain label for grouping
                        domain,
                        original_tld
                    ))
            except Exception as e:
                print(f"Error processing domain {domain}: {e}")

        for future in as_completed(futures):
            result = future.result()
            if result:
                all_data.append(result)

    # ✅ Safety filter (just in case)
    filtered_data = [
        data for data in all_data
        if str(data.get('Domain permutation', '')).lower().endswith(KEEP_TLDS)
    ]

    return filtered_data, permutation_count

def compare_with_previous(df, client_name, today_date):
    previous_files = [f for f in os.listdir() if f.startswith(f'domain_permutations_{client_name}_')]
    previous_files.sort()
    latest_previous_file = previous_files[-1] if previous_files else None

    if latest_previous_file is None:
        print(f"No previous file found for client '{client_name}'.")
        if not df.empty:
            df['New/Changed'] = 'Yes'
        return df

    print(f"Comparing with the latest previous file: '{latest_previous_file}'")
    try:
        previous_df = pd.read_excel(latest_previous_file, sheet_name='Permutations')
    except Exception as e:
        print(f"Couldn't read previous file: {e}")
        return df

    if 'Domain permutation' in previous_df.columns:
        previous_domains = set(previous_df['Domain permutation'])
        previous_status_dict = previous_df.set_index('Domain permutation')['Domain Availability Status'].to_dict()

        def is_new_or_changed(row):
            dp = row['Domain permutation']
            cur_status = str(row.get('Domain Availability Status', '')).strip().lower()
            if dp not in previous_domains:
                return 'Yes'
            prev_status = str(previous_status_dict.get(dp, '')).strip().lower()
            became_active = cur_status.startswith("the website is active") and (
                "parked" in prev_status or "for sale" in prev_status or "offline" in prev_status
            )
            return 'Yes' if became_active else 'No'

        if not df.empty:
            df['New/Changed'] = df.apply(is_new_or_changed, axis=1)
    else:
        print(f"The latest previous file does not contain the 'Domain permutation' column.")

    return df

def main(subdomains, original_tlds=None):
    tlds = ['com', 'net', 'org']  # we still query these bases; permutations are filtered later
    all_data = []
    permutation_summary = []
    colors = {}

    # Color per original base subdomain
    for subdomain in subdomains:
        colors[subdomain] = "{:06x}".format(random.randint(0, 0xFFFFFF))

    for subdomain in subdomains:
        if '.' in subdomain:
            subdomain_name = subdomain.split('.')[0]
        else:
            subdomain_name = subdomain

        original_tld = original_tlds[subdomain]
        print(f"[Base] {subdomain_name}.{original_tld}")

        data, count = process_subdomain(subdomain_name, tlds, original_tld)
        all_data.extend(data)
        permutation_summary.append({
            'Domain': f"{subdomain}.{original_tlds[subdomain]}",
            'Permutations': count
        })

    # Deduplicate by Domain permutation
    unique_map = {}
    for d in all_data:
        dp = d.get('Domain permutation')
        if dp and dp not in unique_map:
            unique_map[dp] = d
    unique_data = list(unique_map.values())

    # Remove original domains from output (just in case)
    unique_data = [data for data in unique_data if f"{data['Original Domain']}" != data['Domain permutation']]

    # DataFrame
    df = pd.DataFrame(unique_data)

    # Compare with previous and mark New/Changed
    today_date = datetime.now().strftime("%Y%m%d")
    if not df.empty:
        df = compare_with_previous(df, client_name, today_date)

    # Summary
    summary_df = pd.DataFrame(permutation_summary)
    if 'Permutations' not in summary_df.columns:
        summary_df['Permutations'] = 0

    # ✅ Robust counts (guard against empty/column missing)
    if df.empty or 'Original Domain' not in df.columns:
        # leave existing counts (from DNSTwister raw counts)
        pass
    else:
        permutation_counts = (
            df['Original Domain']
            .value_counts()
            .rename_axis('Domain')
            .reset_index(name='Permutations')
        )
        summary_df = (
            summary_df.drop(columns=['Permutations'], errors='ignore')
            .merge(permutation_counts, how='left', on='Domain')
        )
        summary_df['Permutations'] = summary_df['Permutations'].fillna(0).astype(int)

    total_permutations = int(summary_df['Permutations'].sum())
    summary_df = pd.concat([summary_df, pd.DataFrame([{'Domain': 'Total', 'Permutations': total_permutations}])], ignore_index=True)

    # Write Excel
    excel_filename = f'domain_permutations_{client_name}_{today_date}.xlsx'
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Permutations', index=False)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    wb = load_workbook(excel_filename)

    for sheet_name in ["Permutations", "Summary"]:

                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # Color header row
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # Optional quality-of-life improvements
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                wb.save(excel_filename)

                # Color 'Permutations' sheet by Original Domain, then remove that column
    try:
        wb = load_workbook(excel_filename)
        ws = wb['Permutations']

        if ws.max_row >= 2:
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = {cell.value: idx for idx, cell in enumerate(header_cells, start=1)}
            orig_col = headers.get('Original Domain')

            if orig_col:
                for r in range(2, ws.max_row + 1):
                    original_domain = ws.cell(row=r, column=orig_col).value or ''
                    sub = str(original_domain).split('.')[0]
                    fill_color = colors.get(sub, "FFFFFF")
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill

                # delete the "Original Domain" column by name
               # ws.delete_cols(orig_col)

        wb.save(excel_filename)
    except Exception as e:
        print(f"Excel post-processing skipped: {e}")

    print(f"Wrote {excel_filename}")

# ---------------- Entry point ----------------
if __name__ == "__main__":
    input_path = DATA_DIR / "company_domains.xlsx"
    if not input_path.exists():
        raise FileNotFoundError(f"Missing input Excel: {input_path}")

    # Expect headers: client, domains  (domains like: "acme.com, beta.org, foo.net")
    df = pd.read_excel(input_path)
    df.columns = [str(c).strip().lower() for c in df.columns]
    if not {"client", "domains"}.issubset(df.columns):
        raise ValueError("Input must have columns: 'client' and 'domains' (comma-separated domains)")

    # Drop blank rows
    df = df.dropna(subset=["client", "domains"])

    for _, row in df.iterrows():
        client = str(row["client"]).strip()
        raw_list = str(row["domains"])

        # Split and normalize each domain
        candidates = [d.strip() for d in raw_list.split(",") if d.strip()]
        cleaned = []
        for d in candidates:
            s = d.lower().strip()
            if s.startswith(("http://", "https://")):
                s = s.split("://", 1)[1]
            if s.startswith("www."):
                s = s[4:]
            s = s.strip(" .")
            if "." in s:
                cleaned.append(s)

        # Build subdomain -> tld (e.g., "acme.com" -> {"acme": "com"})
        # NOTE: if the same subdomain appears with multiple TLDs (e.g., acme.com + acme.org),
        # the last one wins. If you want to handle multiple TLDs per subdomain, tell me and
        # I’ll expand main() to iterate per (subdomain, tld) pair.
        original_tlds = {}
        bad = []
        for s in cleaned:
            parts = s.split(".", 1)
            if len(parts) == 2 and parts[0] and parts[1]:
                sub, tld = parts[0].strip(), parts[1].strip()
                original_tlds[sub] = tld
            else:
                bad.append(s)

        if bad:
            print(f"[{client}] Skipping invalid domains (expected like 'acme.com'): {bad}")

        if not original_tlds:
            print(f"[{client}] No valid domains found. Skipping.")
            continue

        # Set filename prefix via the global used in main()
        client_name = client

        subdomains = list(original_tlds.keys())
        print(f"[Client] {client} → bases: {', '.join(f'{k}.{v}' for k, v in original_tlds.items())}")
        try:
            main(subdomains, original_tlds)
        except Exception as e:
            print(f"[{client}] Error: {e}")

