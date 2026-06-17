# darkweb_monitor.py
import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from requests.exceptions import RequestException
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

load_dotenv()
# ---------- Config ----------

CWD = Path(__file__).resolve().parent.parent
DATA_DIR = (CWD / "data").resolve()
INPUT_XLSX = DATA_DIR / "company_emails.xlsx"
OUT_DIR = (CWD / "outputs").resolve()


# HIBP API key (set as ENV var; fallback is a placeholder)
HIBP_API_KEY = os.getenv("HIBP_API_KEY")
HIBP_SLEEP = float(os.getenv("HIBP_SLEEP", "1.7"))

# ---------- Helpers ----------
def check_hibp(email: str, api_key: str):
    """Query HIBP full breach model for an email. Returns list (0+ entries), [] if none, or None on error."""
    url = f"https://haveibeenpwned.com/api/v3/breachedaccount/{email}?truncateResponse=false"
    headers = {
        "hibp-api-key": api_key,
        "Accept": "application/json",
        "User-Agent": "CTI-Checker/1.0",  
    }

    # simple retry for transient 429/5xx
    for attempt in range(3):
        try:
            resp = requests.get(url, headers=headers, timeout=20)
        except RequestException as e:
            if attempt == 2:
                print(f"[HIBP] request error for {email}: {e}")
                return None
            time.sleep(HIBP_SLEEP)
            continue

        if resp.status_code == 200:
            return resp.json()
        if resp.status_code == 404:
            return []  # no breaches
        if resp.status_code in (429, 500, 503):
            time.sleep(HIBP_SLEEP + 0.5)
            continue

        print(f"[HIBP] {email} -> status {resp.status_code}: {resp.text[:200]}")
        return None

    return None

def check_exposedornot(email: str):
    """Query ExposedOrNot breach analytics. Returns breach details list, [] if none, or None on error."""
    url = f"https://api.xposedornot.com/v1/breach-analytics?email={email}"

    try:
        resp = requests.get(url, timeout=20)

        if resp.status_code == 200:
            data = resp.json()

            exposed_breaches = data.get("ExposedBreaches") or {}
            breaches_details = exposed_breaches.get("breaches_details") or []

            return breaches_details

        if resp.status_code == 404:
            return []

        print(f"[ExposedOrNot] {email} -> status {resp.status_code}: {resp.text[:200]}")
        return None

    except RequestException as e:
        print(f"[ExposedOrNot] request error for {email}: {e}")
        return None

def parse_description(html: str):
    """Extract plain text + reference links from HIBP HTML description."""
    soup = BeautifulSoup(html or "", "html.parser")
    text = soup.get_text(" ", strip=True)
    links = "; ".join(a["href"] for a in soup.find_all("a", href=True))
    return text, links


def split_emails(value) -> list[str]:
    """Split a cell value into emails (comma/semicolon separated), clean empties."""
    if pd.isna(value):
        return []
    s = str(value)
    parts = [p.strip() for p in s.replace(";", ",").split(",")]
    return [p for p in parts if p]


def make_unique(items: list[str]) -> list[str]:
    """Case-insensitive dedupe with naive singular/plural normalization."""
    seen = {}
    ordered = []
    for itm in items:
        if not itm:
            continue
        norm = itm.strip().lower().rstrip("s")
        if norm not in seen:
            seen[norm] = itm.strip()
            ordered.append(itm.strip())
    return ordered


# ---------- Load input ----------
if not INPUT_XLSX.exists():
    raise FileNotFoundError(f"Input file not found: {INPUT_XLSX}")

df = pd.read_excel(INPUT_XLSX)

required_cols = ["client", "personal_emails", "corporate_emails"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Missing required columns in {INPUT_XLSX}: {missing}")

current_date = datetime.now().strftime("%Y-%m-%d")
one_year_ago = pd.Timestamp.now() - pd.DateOffset(years=1)

# ---------- Process ----------
for _, row in df.iterrows():
    client_name = str(row["client"]).strip()
    personal_emails = split_emails(row["personal_emails"])
    corporate_emails = split_emails(row["corporate_emails"])

    # process each email for this client
    client_rows = []
    breach_info_rows = []  # detailed, per-breach; later dedup by Name

    print(f"Processing client: {client_name}")

    for email in dict.fromkeys(personal_emails + corporate_emails):  # dedupe, keep order
        print(f"  Checking email: {email}")

        all_leaks_list = []
        data_classes_set = set()
        most_recent_dt = None
        most_recent_str = "No Data"

        hibp_results = check_hibp(email, HIBP_API_KEY)
        time.sleep(HIBP_SLEEP)

        xon_results = check_exposedornot(email)

        # ---------- HIBP results ----------
        if hibp_results is None:
            print(f"    [WARN] HIBP error for {email}; continuing.")
        elif len(hibp_results) == 0:
            print(f"    No HIBP breaches found.")
        else:
            for entry in hibp_results:
                name = entry.get("Name", "").strip()
                breach_date_str = entry.get("BreachDate", "").strip()
                added_date_str = entry.get("AddedDate", "").strip()
                desc_html = entry.get("Description", "")
                classes = entry.get("DataClasses", []) or []

                all_leaks_list.append(f"{name} ({breach_date_str})")
                data_classes_set.update([c.strip() for c in classes if c and isinstance(c, str)])

                desc_text, refs = parse_description(desc_html)
                breach_info_rows.append([name, breach_date_str, added_date_str, desc_text, refs])

                try:
                    bd = datetime.strptime(breach_date_str, "%Y-%m-%d")
                    if most_recent_dt is None or bd > most_recent_dt:
                        most_recent_dt = bd
                        most_recent_str = f"{name}: {breach_date_str}"
                except Exception:
                    pass

        # ---------- ExposedOrNot results ----------
        if xon_results is None:
            print(f"    [WARN] ExposedOrNot error for {email}; continuing.")
        elif len(xon_results) == 0:
            print(f"    No ExposedOrNot breaches found.")
        else:
            for entry in xon_results:
                name = str(entry.get("breach", "")).strip()
                breach_year = str(entry.get("xposed_date", "")).strip()
                added_date_str = str(entry.get("added", "")).strip()
                desc_text = str(entry.get("details", "")).strip()
                refs = str(entry.get("references", "")).strip()
                xposed_data = str(entry.get("xposed_data", "")).strip()

                classes = [c.strip() for c in xposed_data.split(";") if c.strip()]

                display_date = breach_year if breach_year else "No Date"

                all_leaks_list.append(f"{name} ({display_date})")
                data_classes_set.update(classes)

                breach_info_rows.append([name, breach_year, added_date_str, desc_text, refs])

                try:
                    bd = datetime.strptime(breach_year, "%Y")
                    if most_recent_dt is None or bd > most_recent_dt:
                        most_recent_dt = bd
                        most_recent_str = f"{name}: {breach_year}"
                except Exception:
                    pass

        # ---------- Risk scoring ----------
        unique_data_leaked = make_unique(sorted(data_classes_set, key=lambda s: s.lower()))
        has_password = any("password" in d.lower() for d in unique_data_leaked)

        if most_recent_dt and most_recent_dt > one_year_ago:
            risk_score = "High"
        elif has_password:
            risk_score = "Med"
        else:
            risk_score = "Low"

        all_leaks = "; ".join(all_leaks_list) if all_leaks_list else "No Data"
        hist_data = ", ".join(unique_data_leaked) if unique_data_leaked else "No Data"

        client_rows.append([email, all_leaks, most_recent_str, hist_data, risk_score])

    # write client workbook if we gathered any rows
    if client_rows:
        client_df = pd.DataFrame(
            client_rows,
            columns=[
                "email",
                "Data Leak Names/dates",
                "Most Recent Leak Names/Date",
                "Data Historically Leaked",
                "Risk Score",
            ],
        ).fillna("No Data")

        # Dedup breach detail rows by Data Leak Name (keep most recent entry)
        breach_details_df = pd.DataFrame(
            breach_info_rows,
            columns=["Data Leak Name", "Breach Date", "Date Added to Database", "Description", "References"],
        )
        if not breach_details_df.empty:
            # sort by breach date desc then dedup
            def parse_dt(x):

                s = str(x).strip()
                for fmt in ("%Y-%m-%d", "%Y"):
                    try:
                             return datetime.strptime(s, fmt)
                    except Exception:
                            pass
                return datetime.min

            breach_details_df["__bd"] = breach_details_df["Breach Date"].map(parse_dt)
            breach_details_df = (
            breach_details_df.sort_values("__bd", ascending=False)
            .drop(columns="__bd")
            .drop_duplicates(subset=["Data Leak Name"])
        )

        OUT_DIR.mkdir(parents=True, exist_ok=True)

        out_name = f"{client_name}_DarkWebBreaches_{current_date}.xlsx"
        out_path = OUT_DIR / out_name
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            client_df.to_excel(writer, sheet_name="Leaked Data", index=False)
            if not breach_details_df.empty:
                breach_details_df.to_excel(writer, sheet_name="Breach Details", index=False)
        wb = load_workbook(out_path)
        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

    header_font = Font(
            color="FFFFFF",
            bold=True
        )
    for sheet_name in ["Leaked Data", "Breach Details"]:

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

    wb.save(out_path)
    print(f"  Saved: {out_name}")
print("All reports generated.")
